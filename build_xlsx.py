"""Renders car_deal_scout.json into the two-table master workbook.
Table 1 (models scouted) then Table 2 (live inventory) on ONE sheet, as asked.
The Down-payment and Term cells are live inputs: the Monthly column is a real PMT()
formula that re-computes when you change them in Google Sheets."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\AshishKhera\car-deals\car_deal_scout.json"
OUT = r"C:\Users\AshishKhera\car-deals\Car_Deal_Scout_Master.xlsx"

d = json.load(open(SRC, encoding="utf-8"))
meta, T1, T2 = d["meta"], d["table1"], d["table2"]

HDR = PatternFill("solid", fgColor="1F3B4D"); HDRF = Font(bold=True, color="FFFFFF")
INPUT = PatternFill("solid", fgColor="FFF2CC")
QUOTED = PatternFill("solid", fgColor="D6EAD6"); ASSUMED = PatternFill("solid", fgColor="FCE8E6")
thin = Side(style="thin", color="CCCCCC"); BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
MONEY = '"$"#,##0'; MONEY2 = '"$"#,##0.00'

wb = Workbook(); ws = wb.active; ws.title = "Master"


def header(row_idx, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=row_idx, column=i, value=c)
        cell.fill, cell.font = HDR, HDRF
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BOX


# ---------------- title + live inputs ----------------
ws["A1"] = "CAR DEAL SCOUT - MASTER SHEET"; ws["A1"].font = Font(bold=True, size=15)
ws["A2"] = f"Generated {meta['generated']} - Ontario, from postcode M3A 1B2 - live site: {meta['site']}"
ws["A3"] = ("Out-the-door = (advertised + dealer admin + OMVIC $10 + tire levy $30) x 1.13 HST + $150 plates. "
            "Dealer admin is $2,334 where Lexus on the Park quoted it, $495 assumed elsewhere.")
ws["A5"] = "DOWN PAYMENT"; ws["B5"] = meta["down"]; ws["B5"].number_format = MONEY; ws["B5"].fill = INPUT
ws["A6"] = "TERM (months)"; ws["B6"] = meta["term"]; ws["B6"].fill = INPUT
ws["C5"] = "<- change these; the Monthly column recalculates"; ws["C5"].font = Font(italic=True, color="666666")
ws["A5"].font = ws["A6"].font = Font(bold=True)
DOWN_REF, TERM_REF = "$B$5", "$B$6"

# ---------------- TABLE 1 ----------------
r = 8
ws.cell(row=r, column=1, value="TABLE 1 - MODELS SCOUTED  (rate source tells you whether the rate was quoted to you or assumed)").font = Font(bold=True, size=12)
r += 1
cols1 = ["Make", "Model", "Trim", "MSRP", "Freight+PDI+AC", "Advertised", "Dealer admin",
         "OUT THE DOOR", "Rate %", "Rate source", "Term", "Down", "MONTHLY", "Dep %/yr", "L/100 km", "Notes"]
header(r, cols1); t1_head = r
for m in T1:
    r += 1
    vals = [m["make"], m["model"], m["trim"], m["msrp"], m["freight"], m["advertised"], m["admin"],
            m["otd"], m["rate"], m["rateSrc"], None, None, None, m["dep"], m["lp100"], m["note"]]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=v); c.border = BOX
        c.alignment = Alignment(wrap_text=(i in (3, 10, 16)), vertical="top")
    for col in (4, 5, 6, 7, 8):
        ws.cell(row=r, column=col).number_format = MONEY
    if m["otd"] is not None and m["rate"] is not None:
        ws.cell(row=r, column=11, value=f"={TERM_REF}")
        ws.cell(row=r, column=12, value=f"={DOWN_REF}").number_format = MONEY
        # -PMT(rate/12, term, OTD - down)
        ws.cell(row=r, column=13, value=f"=-PMT(I{r}/100/12,{TERM_REF},H{r}-{DOWN_REF})").number_format = MONEY2
        ws.cell(row=r, column=13).font = Font(bold=True)
    src = (m["rateSrc"] or "").upper()
    fill = QUOTED if src.startswith(("QUOTED", "PUBLISHED")) else (ASSUMED if src else None)
    if fill:
        ws.cell(row=r, column=9).fill = fill; ws.cell(row=r, column=10).fill = fill
t1_end = r

# ---------------- TABLE 2 ----------------
r += 3
ws.cell(row=r, column=1, value=f"TABLE 2 - LIVE INVENTORY  ({len(T2)} vehicles, every link verified this week)").font = Font(bold=True, size=12)
r += 1
cols2 = ["Make", "Model", "Trim", "Year", "Asking price", "MSRP", "MSRP source", "Odometer km",
         "Asking vs MSRP", "Status / notes", "Seller", "Listing", "Scan"]
header(r, cols2); t2_head = r
for v in T2:
    r += 1
    vals = [v["make"], v["model"], v["trim"], v["year"], v["price"], v["msrp"], v["msrpSrc"], v["km"],
            None, v["status"], v["seller"], None, v["src"]]
    for i, val in enumerate(vals, 1):
        c = ws.cell(row=r, column=i, value=val); c.border = BOX
        c.alignment = Alignment(wrap_text=(i in (3, 7, 10)), vertical="top")
    ws.cell(row=r, column=5).number_format = MONEY
    ws.cell(row=r, column=6).number_format = MONEY
    ws.cell(row=r, column=8).number_format = '#,##0'
    if v["msrp"]:
        ws.cell(row=r, column=9, value=f"=E{r}-F{r}").number_format = '"$"#,##0;[Red]-"$"#,##0'
    url = v["url"].replace('"', '%22')
    ws.cell(row=r, column=12, value=f'=HYPERLINK("{url}","open")').font = Font(color="0563C1", underline="single")

for i, w in enumerate([10, 16, 30, 9, 12, 12, 12, 13, 12, 44, 10, 11, 12, 9, 9, 60], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A8"

# ---------------- Notes sheet ----------------
n = wb.create_sheet("Read me")
notes = [
    ("What Table 1 is", "One row per model/trim we scouted. MSRP is the manufacturer sticker ex-freight; Advertised is the number a dealer actually prints; OUT THE DOOR adds fees + 13% HST + plates."),
    ("How to read Rate source", "GREEN = a rate that was quoted to you or published by the dealer/manufacturer. PINK = assumed. Where nothing was obtained I used 6.99% so the payment column is filled - treat those payments as placeholders and get a real quote."),
    ("The monthly payment", "-PMT(rate/12, term, OTD - down). Change B5 (down) or B6 (term) and every Monthly cell recalculates. Default $15,000 down / 60 months."),
    ("Lexus rates are the real ones", "2.9%/60 finance and 1.9%/48 lease were written on Lexus on the Park worksheets on 2026-09-02 as 'special rate'. The 5.5% on the 500h is Don Valley North's published rate."),
    ("The freight-twice catch", "Lexus on the Park's finance sheet used $73,258 as the sell price - that is Lexus Canada's 'from' price, which already includes freight - then added $870 freight again plus $2,334 dealer fees. Their $86,611 OTD is about $2.7K above the model here. Worth asking about."),
    ("What Table 2 is", "Every live listing from this week's scans (Sept 1-2), one row each, with its own link. Older scans (the early CR-V / RAV4 / Grand Highlander gas sweep) are NOT included because some of those cars have since sold."),
    ("MSRP on used cars", "Where the exact original sticker is not on the listing, the MSRP column shows the same-trim sticker (2021 launch pricing for Genesis and Palisade, current-MY equivalent for Lexus) and the source column says which. 'Asking vs MSRP' is negative when the car is below sticker."),
    ("Your standing inputs", "29,400 km/yr (120 km x 245 weekdays). Lexus basic warranty 4yr/80,000km, Certified 6yr/120,000km; Genesis and Hyundai 5yr/100,000km; Mitsubishi 10yr/160,000km powertrain. At your mileage the km cap always binds first."),
]
n.append(["Topic", "Detail"]); n["A1"].fill = n["B1"].fill = HDR; n["A1"].font = n["B1"].font = HDRF
for a, b in notes:
    n.append([a, b])
n.column_dimensions["A"].width = 28; n.column_dimensions["B"].width = 120
for row in n.iter_rows(min_row=2, max_row=1 + len(notes)):
    for c in row:
        c.alignment = Alignment(wrap_text=True, vertical="top"); c.border = BOX

wb.save(OUT)
print("saved", OUT)
print(f"  Table 1 rows {t1_head + 1}-{t1_end} ({t1_end - t1_head} models) | Table 2 rows {t2_head + 1}-{r} ({r - t2_head} listings)")
